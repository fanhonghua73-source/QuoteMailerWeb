"""
Excel 解析与图片提取模块 v3
统一字段名：小写下划线
"""

import os
import shutil
import tempfile
import uuid
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet


class ExcelParser:
    """Excel 报价单解析器"""
    
    def __init__(self, excel_path: str, output_dir: str = None):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path, data_only=True)
        
        # 生成全局唯一的目录后缀，防止多线程并发时临时文件互相覆盖或误删
        unique_session_id = uuid.uuid4().hex
        
        if output_dir:
            self.output_dir = Path(output_dir) / f".temp_assets_{unique_session_id}"
        else:
            self.output_dir = Path(excel_path).parent / f".temp_assets_{unique_session_id}"
        
        self.temp_dir = self.output_dir / "images"
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        
    def parse(self, sheet_index: int = 0) -> Dict:
        """解析 Excel，返回标准化字典列表"""
        sheet_name = self.wb.sheetnames[sheet_index]
        ws = self.wb[sheet_name]
        
        # 提取表头
        headers = [cell.value for cell in ws[1]]
        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(headers)]
        
        # 构建图片索引
        row_images = self._build_image_index(ws)
        
        # 提取数据行
        products = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
                
            row_data = dict(zip(headers, row))
            
            # 标准化字段
            product = self._normalize_product(row_data)
            
            # 绑定图片
            product['images'] = row_images.get(row_idx, [])
            
            products.append(product)
        
        # 打印日志
        print(f"[ExcelParser] 解析完成: {len(products)} 个产品")
        if products:
            print(f"[ExcelParser] 第一条数据: {products[0]}")
        
        return {
            'headers': headers,
            'products': products,
            'assets_dir': str(self.temp_dir)
        }
    
    def _build_image_index(self, ws: Worksheet) -> Dict[int, List[str]]:
        """构建行号到图片的索引，按列排序"""
        row_images = {}
        image_positions = []  # [(row, col, path), ...]
        
        if not hasattr(ws, '_images') or not ws._images:
            return row_images
        
        for idx, img in enumerate(ws._images):
            try:
                anchor = img.anchor
                row = anchor._from.row + 1  # Excel行号从1开始
                col = anchor._from.col + 1  # Excel列号从1开始
                
                filename = self._save_image(img, idx, row)
                if filename:
                    abs_path = os.path.abspath(str(self.temp_dir / filename))
                    image_positions.append((row, col, abs_path))
            except Exception as e:
                print(f"Warning: 跳过图片 {idx}: {e}")
                continue
        
        # 按行分组，按列排序
        image_positions.sort(key=lambda x: (x[0], x[1]))  # 先按行，再按列排序
        
        for row, col, path in image_positions:
            if row not in row_images:
                row_images[row] = []
            row_images[row].append(path)
        
        print(f"[ExcelParser] 图片索引: {row_images}")
        return row_images
    
    def _save_image(self, img: XLImage, idx: int, row: int) -> Optional[str]:
        """保存图片"""
        try:
            img_data = img._data()
            ext = self._detect_image_extension(img_data)
            filename = f"product_{row}_{idx}.{ext}"
            filepath = self.temp_dir / filename
            filepath.write_bytes(img_data)
            return filename
        except Exception as e:
            print(f"Error saving image: {e}")
            return None
    
    def _detect_image_extension(self, data: bytes) -> str:
        if data[:4] == b'\x89PNG':
            return 'png'
        elif data[:3] == b'\xFF\xD8\xFF':
            return 'jpg'
        elif data[:4] == b'GIF8':
            return 'gif'
        else:
            return 'png'
    
    def _normalize_product(self, row_data: Dict) -> Dict:
        """标准化字段名 - 统一使用小写下划线"""
        # 字段映射表
        field_map = {
            'item': ['item', 'item_no', 'item_no', '品名', '型号', 'Item'],
            'name': ['name', 'product_name', '产品名称', 'Name', '品名'],
            'spec': ['spec', 'specification', '规格', '型号', 'Specification'],
            'packing': ['packing', '包装', 'Packing'],
            'cbm': ['cbm', '体积', 'CBM', 'cuft'],
            'price': ['price', 'price_fob', '单价', '价格', 'Price', 'Price FOB'],
            'moq': ['moq', '最小起订量', 'MOQ', 'moq'],
            'material': ['material', '材质', 'Material'],
            'size': ['size', '尺寸', 'Size'],
        }
        
        product = {}
        
        for std_name, possible_names in field_map.items():
            value = ''
            for name in possible_names:
                if name in row_data and row_data[name]:
                    value = row_data[name]
                    break
            product[std_name] = value
        
        # 处理价格
        if product.get('price'):
            price_str = str(product['price']).replace('$', '').replace('¥', '').replace(',', '').strip()
            try:
                product['price'] = float(price_str)
            except:
                product['price'] = price_str
        
        return product
    
    def cleanup(self):
        """清理临时文件"""
        if self.output_dir.exists():
            shutil.rmtree(self.output_dir)


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        parser = ExcelParser(sys.argv[1])
        result = parser.parse()
        print(f"\n最终输出: {result['products']}")
